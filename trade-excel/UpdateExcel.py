from typing import List, Tuple
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
import JustEtfScrape as etf_scrape
import subprocess
import tkinter as tk
from tkinter import filedialog
import json
import os

JSON_FILE_PATH = "./config.json"
EXCEL_PATH = "E:/Documents/Trade/Trade.xlsx"  # TODO change to receive as input
SHEET_NAME = (
    "justETF Data"  # TODO change to receive as input, also change to investigate
)
TICKER_TITLE = "Isin"  # TODO change to receive as input
COLUMN_WIDTH_PADDING = 3
HEADER_COLOR = PatternFill(start_color="C2FFC2", end_color="C2FFC2", fill_type="solid")
BREAKLINE = "\n"

JSON_PATHS = "Paths"
JSON_BROWSER = "Browser"

COLUMNS = [
    "Isin",
    "Name",
    "Tickers",
    "Currency",
    "Volatility 1 year",
    "Volatility 3 years",
    "Volatility 5 years",
    "Returns 1 year",
    "Returns 3 years",
    "Returns 5 years",
    "TER",
    "Distribution",
    "Replication",
    "Countries",
    "Sectors",
    "Fund size",
    "Number of holdings",
]


def FlattenListToString(list: List, delimiter: str) -> str:
    return delimiter.join(list)


def LoadEtfISINFromExcel(file_path: str, column_name) -> List[str]:
    column_values = pd.read_excel(
        file_path, sheet_name=SHEET_NAME, usecols=[column_name], header=0
    )[column_name].tolist()
    return column_values


def GetEtfISINInformation(etfs_isins: List[str], browser_path: str) -> pd.DataFrame:
    etfs_data = etf_scrape.SeleniumScrape(etfs_isins, browser_path)

    etf_dataframe = pd.DataFrame(columns=COLUMNS)
    print(etf_dataframe)

    for isin in etf_isins:
        if not pd.isna(isin):
            etf_data = etfs_data[isin]
            row_data = [
                isin,
                etf_data[etf_scrape.DICT_NAME],
                FlattenListToString(etf_data[etf_scrape.DICT_TICKERS], ", "),
                etf_data[etf_scrape.DICT_CURRENCY],
                etf_data[etf_scrape.DICT_VOLATILITY][0],
                etf_data[etf_scrape.DICT_VOLATILITY][1],
                etf_data[etf_scrape.DICT_VOLATILITY][2],
                etf_data[etf_scrape.DICT_RETURNS][0],
                etf_data[etf_scrape.DICT_RETURNS][1],
                etf_data[etf_scrape.DICT_RETURNS][2],
                etf_data[etf_scrape.DICT_TER],
                etf_data[etf_scrape.DICT_DISTRIBUTION],
                etf_data[etf_scrape.DICT_REPLICATION],
                FlattenListToString(etf_data[etf_scrape.DICT_COUNTRIES], BREAKLINE),
                FlattenListToString(etf_data[etf_scrape.DICT_SECTORS], BREAKLINE),
                etf_data[etf_scrape.DICT_FUND_SIZE],
                etf_data[etf_scrape.DICT_NUM_HOLDINGS],
            ]
            new_row = pd.DataFrame([row_data], columns=COLUMNS)
            print(new_row)
            etf_dataframe = pd.concat([etf_dataframe, new_row], ignore_index=True)

            print(etf_dataframe)
    return etf_dataframe


def WriteToExcel(etfs_info: pd.DataFrame):
    with pd.ExcelWriter(
        path=EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as excel_writer:
        etfs_info.to_excel(
            excel_writer, sheet_name=SHEET_NAME, index=False, header=True
        )
        worksheet = excel_writer.sheets[SHEET_NAME]
        # Auto-fit columns and center align
        for i, column in enumerate(etfs_info.columns):
            column_len = max(etfs_info[column].astype(str).map(len).max(), len(column))
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=i + 1).column_letter
            ].width = str(
                (int(column_len) + COLUMN_WIDTH_PADDING)
                / (2 if (column == COLUMNS[13] or column == COLUMNS[14]) else 1)
            )  # Reduce padding if it is Countries of Sectors columns

            for cell in worksheet[worksheet.cell(row=1, column=i + 1).column_letter]:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        # Make the header row thicker
        worksheet.row_dimensions[
            1
        ].height = 30  # Adjust as needed for the desired height
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_COLOR


def ReadPathFromJson(path_field: str) -> str:
    # Check if the JSON file exists
    if not os.path.exists(JSON_FILE_PATH):
        print("JSON file does not exist.")
        return ""

    with open(JSON_FILE_PATH, "r") as file:
        try:
            data = json.load(file)
        except json.JSONDecodeError:
            print("Error decoding JSON.")
            return ""

    # Check if the JSON is not empty and contains the required field
    if data and (JSON_PATHS in data) and (JSON_BROWSER in data[JSON_PATHS]):
        return data[JSON_PATHS][path_field]
    else:
        print("Required field 'ExePaths' or 'Browser' not found in JSON.")
        return ""


def WritePathToJson(browser_path: str):
    # Initialize the data structure
    data = {}

    # Check if the JSON file exists and is not empty
    if os.path.exists(JSON_FILE_PATH):
        with open(JSON_FILE_PATH, "r") as file:
            try:
                data = json.load(file)
            except json.JSONDecodeError:
                print("Error decoding JSON. The file might be empty or corrupted.")
                data = {}

    # Ensure the necessary structure exists
    if JSON_PATHS not in data:
        data[JSON_PATHS] = {}

    # Update the Browser path
    data[JSON_PATHS][JSON_BROWSER] = browser_path

    # Write the updated data back to the JSON file
    with open(JSON_FILE_PATH, "w") as file:
        json.dump(data, file, indent=4)
    print(f"Browser path '{browser_path}' has been written to the JSON file.")


def UpdateExcelPopup(initial_path: str = "") -> Tuple[str, bool, bool]:
    def browse_file():
        filepath = filedialog.askopenfilename()
        if filepath:
            entry_path.delete(0, tk.END)  # Clear the entry widget
            entry_path.insert(0, filepath)  # Insert the selected file path

    def on_ok():
        result_dict["filepath"] = entry_path.get()
        result_dict["update_excel"] = var_update_excel.get()
        result_dict["open_excel"] = True
        root.destroy()

    def on_close():
        result_dict["filepath"] = initial_path
        result_dict["update_excel"] = False
        result_dict["open_excel"] = False
        root.destroy()

    result_dict = {}

    # Create a root window
    root = tk.Tk()
    root.title("Selector")

    # Set up the grid layout
    root.columnconfigure(0, weight=1)
    root.columnconfigure(1, weight=3)
    root.columnconfigure(2, weight=1)

    # Add a label and entry for the file path
    label_path = tk.Label(root, text="Chromium based browser filepath:")
    label_path.grid(column=0, row=0, padx=10, pady=10, sticky=tk.W)

    entry_path = tk.Entry(root, width=50)
    entry_path.grid(column=1, row=0, padx=10, pady=10, sticky=tk.W)
    entry_path.insert(
        0, initial_path
    )  # Pre-fill the entry widget with the initial path

    browse_button = tk.Button(root, text="Browse", command=browse_file)
    browse_button.grid(column=2, row=0, padx=10, pady=10, sticky=tk.W)

    # Add a checkbox to ask if the Excel file should be updated
    var_update_excel = tk.BooleanVar()
    checkbox_update_excel = tk.Checkbutton(
        root, text="Update Excel?", variable=var_update_excel
    )
    checkbox_update_excel.grid(column=0, row=1, padx=10, pady=10, sticky=tk.W)

    # Add OK button
    ok_button = tk.Button(root, text="OK", command=on_ok)
    ok_button.grid(column=2, row=2, padx=10, pady=10, sticky=tk.E)

    # Handle window close event
    root.protocol("WM_DELETE_WINDOW", on_close)

    # Center the window on the screen
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    # Run the Tkinter main loop
    root.mainloop()

    return (
        result_dict["filepath"],
        result_dict["update_excel"],
        result_dict["open_excel"],
    )


if __name__ == "__main__":
    browser_path = ReadPathFromJson(JSON_BROWSER)

    browser_path_final, update_excel, open_excel = UpdateExcelPopup(browser_path)

    if not open_excel:
        exit()

    if browser_path_final != browser_path:
        WritePathToJson(browser_path_final)

    if update_excel:
        etf_isins = LoadEtfISINFromExcel(EXCEL_PATH, TICKER_TITLE)

        etfs_info = GetEtfISINInformation(etf_isins, browser_path_final)
        WriteToExcel(etfs_info)

    with subprocess.Popen(["start", "/WAIT", EXCEL_PATH], shell=True) as doc:
        doc.wait()  # Waits until the user closes Excel
