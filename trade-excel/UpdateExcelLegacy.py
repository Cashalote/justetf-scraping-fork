import re
from typing import List
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
import tkinter as tk
from tkinter import messagebox
import os

import justetf_scraping.overview as justetf

CSV_PATH = "E:\\Documents\\Trade\\etf_data.csv"
EXCEL_PATH = "E:\\Documents\\Trade\\Test.xlsx"  # TODO change to receive as input
SHEET_NAME = (
    "justETF Data"  # TODO change to receive as input, also change to investigate
)
TICKER_TITLE = "ETF Ticker"  # TODO change to receive as input
COLUMN_WIDTH_PADDING = 2
HEADER_COLOR = PatternFill(start_color="C2FFC2", end_color="C2FFC2", fill_type="solid")

FILTER_COLUMNS = [
    # IDs
    "isin",
    "wkn",
    "ticker",
    "valor",
    # Basic info
    "name",
    # "groupValue": "index",
    "inception_date",
    "strategy",  # Custom field added during request
    "domicile_country",
    "currency",
    "securities_lending",
    "distribution",
    "ter",
    "replication",
    "fund_size",
    "is_sustainable",
    "number_of_holdings",
    # Value return
    "yesterday",
    "last_week",
    "last_month",
    "last_three_months",
    "last_six_months",
    "last_year",
    "last_three_years",
    "last_five_years",
    # Dividends
    "last_dividends",
    "last_year_dividends",
    # Volatility
    "last_year_volatility",
    "last_three_years_volatility",
    "last_five_years_volatility",
    # Return/Risk
    "last_year_return_per_risk",
    "last_three_years_return_per_risk",
    "last_five_years_return_per_risk",
    # Drawdown
    "max_drawdown",
    "last_year_max_drawdown",
    "last_three_years_max_drawdown",
    "last_five_years_max_drawdown",
]


def SaveData(dataframe: pd.DataFrame):
    dataframe.to_csv(CSV_PATH, index=False, mode="w")


def LoadData() -> pd.DataFrame:
    return pd.read_csv(CSV_PATH)


def RequestETFData() -> pd.DataFrame:
    etfs_de = justetf.load_overview(local_country="DE", enrich=True).reset_index()
    etfs_at = justetf.load_overview(local_country="AT", enrich=True).reset_index()
    etfs_gb = justetf.load_overview(local_country="GB", enrich=True).reset_index()
    etfs_it = justetf.load_overview(local_country="IT", enrich=True).reset_index()
    etfs_fr = justetf.load_overview(local_country="FR", enrich=True).reset_index()
    etfs_es = justetf.load_overview(local_country="ES", enrich=True).reset_index()
    etfs_nl = justetf.load_overview(local_country="NL", enrich=True).reset_index()
    etfs_be = justetf.load_overview(local_country="BE", enrich=True).reset_index()

    # Use etfs_de as the base DataFrame
    etfs = etfs_de.copy()

    # Append new lines from etfs_gb where ticker is different
    etfs = pd.concat(
        [etfs, etfs_at[~etfs_at["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_nl where ticker is different
    etfs = pd.concat(
        [etfs, etfs_gb[~etfs_gb["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_gb where ticker is different
    etfs = pd.concat(
        [etfs, etfs_it[~etfs_it["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_nl where ticker is different
    etfs = pd.concat(
        [etfs, etfs_fr[~etfs_fr["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_gb where ticker is different
    etfs = pd.concat(
        [etfs, etfs_es[~etfs_es["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_nl where ticker is different
    etfs = pd.concat(
        [etfs, etfs_nl[~etfs_nl["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    # Append new lines from etfs_gb where ticker is different
    etfs = pd.concat(
        [etfs, etfs_be[~etfs_be["ticker"].isin(etfs["ticker"])]], ignore_index=True
    )

    return etfs


def UpdateCSVPopup() -> bool:
    # Create a root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Ask the yes/no question
    result = messagebox.askyesno("Update CSV", "Do you wish to update the CSV file?")

    # Destroy the root window
    root.destroy()

    # Handle the response
    return result


def SetupData() -> pd.DataFrame:
    etfs = pd.DataFrame()
    try:
        if not os.path.exists(CSV_PATH):
            etfs = RequestETFData()
            SaveData(etfs)
            return etfs
        etf_data = pd.read_csv(CSV_PATH)
        if etf_data.empty:
            etfs = RequestETFData()
            SaveData(etfs)
        else:
            if UpdateCSVPopup():
                etfs = RequestETFData()
                SaveData(etfs)
            else:
                etfs = LoadData()
    except pd.errors.EmptyDataError:
        etfs = RequestETFData()
        SaveData(etfs)
    return etfs


def LoadEtfTickersFromExcel(file_path: str, column_name) -> List[str]:
    column_values = pd.read_excel(
        file_path, sheet_name=SHEET_NAME, usecols=[column_name], header=0
    )[column_name].tolist()
    return column_values


def CleanTickers(etf_tickers: List[str]) -> List[str]:
    # Remove suffix from ETF ticker (.DE, .UK, etc.)
    return [re.sub(r"\..*", "", ticker) for ticker in etf_tickers]


def GetTickerInformation(etf_df: pd.DataFrame, etf_tickers: List[str]) -> pd.DataFrame:
    cleaned_tickers = CleanTickers(etf_tickers)

    # Ensure the column names are stripped of any leading/trailing whitespace
    etf_df.columns = etf_df.columns.str.strip()

    # Filter ETFs based on the cleaned tickers
    filtered_etfs = etf_df.loc[etf_df["ticker"].isin(cleaned_tickers)]

    # Create dataframe with cleaned ticker in column to make merge
    excel_ticker_column = pd.DataFrame(cleaned_tickers, columns=[TICKER_TITLE])

    # Concatenate existing excel data with new data
    combined_data = pd.merge(
        excel_ticker_column,
        filtered_etfs,
        left_on=TICKER_TITLE,
        right_on="ticker",
        how="outer",
    )

    # Use the original ticker values
    combined_data[TICKER_TITLE] = etf_tickers

    return combined_data


def WriteToExcel(etfs_info: pd.DataFrame):
    with pd.ExcelWriter(
        path=EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as excel_writer:
        etfs_info.to_excel(
            excel_writer, sheet_name=SHEET_NAME, index=False, header=True
        )
        worksheet = excel_writer.sheets[SHEET_NAME]
        # Auto-fit columns and center align
        for i, column in enumerate(etfs_info.columns):
            column_len = max(etfs_info[column].astype(str).map(len).max(), len(column))
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=i + 1).column_letter
            ].width = str(int(column_len) + COLUMN_WIDTH_PADDING)

            for cell in worksheet[worksheet.cell(row=1, column=i + 1).column_letter]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Make the header row thicker
        worksheet.row_dimensions[
            1
        ].height = 30  # Adjust as needed for the desired height
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_COLOR


if __name__ == "__main__":
    etfs = SetupData()

    filtered_etfs = etfs[FILTER_COLUMNS]

    etf_tickers = LoadEtfTickersFromExcel(EXCEL_PATH, TICKER_TITLE)
    etfs_info = GetTickerInformation(filtered_etfs, etf_tickers)
    etfs_info_sorted = etfs_info.sort_values(by=etfs_info.columns[0], ascending=True)
    WriteToExcel(etfs_info_sorted)
